"""Audit log: records every accept/deny/auto_accept decision.

Entries are appended to JSON-lines files in logs/audit/YYYY-WNN.jsonl
(one file per ISO week). A weekly Excel export (openpyxl) is generated
at daemon startup for any week that has a .jsonl but no .xlsx yet.
"""
from __future__ import annotations

import json
import logging
import os
import threading
from collections import Counter
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)


@dataclass
class AuditEntry:
    timestamp: str
    week: str
    request_id: str
    connector: str
    tool: str
    tool_name: str
    summary: str
    sender: str
    decision: str           # "approved" | "rejected" | "auto_accepted"
    auto_accept_rule: str   # rule name if auto_accepted, else ""
    latency_seconds: float


class AuditLogger:
    def __init__(self, log_dir: str) -> None:
        self._log_dir = Path(log_dir)
        self._log_dir.mkdir(parents=True, exist_ok=True)
        self._lock = threading.Lock()

    def record(self, entry: AuditEntry) -> None:
        week_file = self._log_dir / f"{entry.week}.jsonl"
        line = json.dumps(asdict(entry)) + "\n"
        with self._lock:
            with open(week_file, "a", encoding="utf-8") as fh:
                fh.write(line)
        logger.debug("Audit: %s %s/%s", entry.decision, entry.connector, entry.tool)

    def export_week_to_excel(self, week: str) -> Optional[str]:
        """Export one week's .jsonl to .xlsx. Returns output path or None."""
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.warning("openpyxl not installed — skipping Excel audit export")
            return None

        week_file = self._log_dir / f"{week}.jsonl"
        if not week_file.exists():
            return None

        entries: list[AuditEntry] = []
        with open(week_file, encoding="utf-8") as fh:
            for line in fh:
                line = line.strip()
                if line:
                    try:
                        entries.append(AuditEntry(**json.loads(line)))
                    except Exception:
                        pass
        if not entries:
            return None

        output_path = str(self._log_dir / f"{week}.xlsx")
        wb = openpyxl.Workbook()

        # ── Main sheet ────────────────────────────────────────────────────
        ws = wb.active
        ws.title = "Decisions"

        HEADERS = [
            "Timestamp", "Week", "Connector", "Tool", "Human-Readable Name",
            "Summary", "Sender / Context", "Decision", "Auto-Accept Rule", "Latency (s)"
        ]
        COL_WIDTHS = [22, 10, 12, 30, 22, 55, 30, 14, 22, 12]

        hdr_font  = Font(bold=True, color="FFFFFF")
        hdr_fill  = PatternFill("solid", fgColor="2D4A6B")
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        decision_fills = {
            "approved":     PatternFill("solid", fgColor="E8F5E9"),
            "auto_accepted": PatternFill("solid", fgColor="E3F2FD"),
            "rejected":     PatternFill("solid", fgColor="FFEBEE"),
        }

        ws.append(HEADERS)
        for col, _ in enumerate(HEADERS, 1):
            c = ws.cell(row=1, column=col)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = hdr_align

        for entry in entries:
            ws.append([
                entry.timestamp, entry.week, entry.connector, entry.tool,
                entry.tool_name, entry.summary, entry.sender, entry.decision,
                entry.auto_accept_rule or "", round(entry.latency_seconds, 2),
            ])
            fill = decision_fills.get(entry.decision, PatternFill())
            for col in range(1, len(HEADERS) + 1):
                ws.cell(row=ws.max_row, column=col).fill = fill

        for col, width in enumerate(COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

        # ── Summary sheet ─────────────────────────────────────────────────
        ws2 = wb.create_sheet("Summary")
        ws2.append(["Metric", "Value"])
        ws2.append(["Week", week])
        ws2.append(["Total decisions", len(entries)])
        counts = Counter(e.decision for e in entries)
        ws2.append(["Approved (manual)", counts.get("approved", 0)])
        ws2.append(["Auto-accepted", counts.get("auto_accepted", 0)])
        ws2.append(["Rejected", counts.get("rejected", 0)])
        ws2.append([])
        ws2.append(["By connector", ""])
        for connector, cnt in sorted(Counter(e.connector for e in entries).items()):
            ws2.append([connector, cnt])
        ws2.column_dimensions["A"].width = 24
        ws2.column_dimensions["B"].width = 14

        wb.save(output_path)
        logger.info("Audit Excel exported: %s (%d entries)", output_path, len(entries))
        return output_path

    def export_all_pending(self) -> None:
        """Export any week that has .jsonl but no .xlsx."""
        for jsonl in sorted(self._log_dir.glob("*.jsonl")):
            week = jsonl.stem
            xlsx = self._log_dir / f"{week}.xlsx"
            if not xlsx.exists():
                self.export_week_to_excel(week)


def current_week() -> str:
    iso = datetime.now(timezone.utc).isocalendar()
    return f"{iso[0]}-W{iso[1]:02d}"


_INSTANCE: Optional[AuditLogger] = None
_LOCK = threading.Lock()


def get_audit_logger() -> AuditLogger:
    global _INSTANCE
    if _INSTANCE is None:
        with _LOCK:
            if _INSTANCE is None:
                fallback = os.path.join(os.path.expanduser("~"), ".loopline", "audit")
                _INSTANCE = AuditLogger(fallback)
    return _INSTANCE


def init_audit_logger(log_dir: str) -> AuditLogger:
    global _INSTANCE
    with _LOCK:
        _INSTANCE = AuditLogger(log_dir)
    return _INSTANCE
