"""Audit log: records every accept/deny/auto_accept decision.

Entries are appended to JSON-lines files in logs/audit/YYYY-WNN.jsonl
(one file per ISO week). A weekly Excel export (openpyxl) is generated
at daemon startup for any week that has a .jsonl but no .xlsx yet.
"""
from __future__ import annotations

import json
import logging
import os
import threading
from collections import Counter
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path

logger = logging.getLogger(__name__)

# Decisions where the AI actually received the data or the write went
# through. Used by AuditLogger.recent_matches() below to count how many
# times a request has already been let through, not merely asked about --
# and by gate.py to decide whether an approved-request's PII match ever got
# released at all, before it will record the literal (or redacted) matched
# text in pii_match_details rather than the "hidden" placeholder. Public
# (no leading underscore) for that second, cross-module use.
APPROVED_LIKE_DECISIONS = frozenset({
    "approved", "auto_accepted", "accepted_via_accept_all", "accepted_via_temp_session",
})


@dataclass
class AuditEntry:
    timestamp: str
    week: str
    request_id: str
    connector: str
    tool: str
    tool_name: str
    summary: str
    sender: str
    decision: str           # "approved" | "rejected" | "auto_accepted" | "accepted_via_accept_all" |
                            # "accepted_via_temp_session" | "denied_unattended" | "policy_check" |
                            # "rules_listed" |
                            # "unattended_session_started" | "unattended_session_ended" |
                            # "rule_changed_via_bridge_proposal" | "rule_removed_via_bridge_proposal" |
                            # "grant_changed_via_bridge_proposal" | "grant_removed_via_bridge_proposal" |
                            # "bridge_proposal_no_op" | "error"
                            # ("error": gate.py's gated_call exited without reaching a normal decision
                            #  branch -- a fallback so an unanticipated failure still leaves a trail)
                            # ("denied_unattended": gate.py denied the call without ever prompting,
                            #  because the connection was in an unattended session and no auto-accept
                            #  rule matched -- distinct from "rejected", which is a human's own Deny.
                            #  Also used by gate.py's propose_rule_change() for the same reason)
                            # ("policy_check": ipc_server.py's check_policy handler -- a preflight
                            #  question, not a real decision; recorded for pattern-spotting only)
                            # ("rules_listed": ipc_server.py's list_rules handler -- not a decision
                            #  either, but the full current rule/grant set was disclosed, worth its
                            #  own record for the same pattern-spotting reason as "policy_check")
                            # ("unattended_session_started"/"_ended": ipc_server.py's begin/end_
                            #  unattended_session handlers, and the same on disconnect cleanup --
                            #  this connection's gate posture changed, which is worth a record of
                            #  its own even though no specific tool call was involved)
                            # ("rule_changed_via_bridge_proposal"/"rule_removed_via_bridge_proposal"/
                            #  "grant_changed_via_bridge_proposal"/"grant_removed_via_bridge_proposal":
                            #  gate.py's propose_rule_change() -- a bridge-initiated auto_accept_rules/
                            #  auto_accept_grants edit that a human confirmed via the same
                            #  show_rule_confirmation_popup() the "Always allow" flow uses, and that
                            #  actually changed something (config's own `changed` return value was
                            #  True). "rejected" is reused, not a new value, when the human declines
                            #  instead)
                            # ("bridge_proposal_no_op": same propose_rule_change() confirmation flow,
                            #  but the human's "yes" didn't actually change anything -- e.g. Claude
                            #  proposed removing a rule/grant value that was already gone. Distinct
                            #  from "rejected" (the human said no) and from the four decisions above
                            #  (a real change happened) -- confirmed and yet a no-op is its own case)
    auto_accept_rule: str   # rule name if auto_accepted, else ""
    latency_seconds: float
    pii_detected: bool = False  # True if pii_detector.py flagged the content before this decision
    pii_categories: list[str] = field(default_factory=list)  # Which category label(s) pii_detector.py
                              # flagged (e.g. "IBAN (bank account number)") -- always populated
                              # whenever pii_detected is True, regardless of the audit_match_details
                              # trial setting below. Category labels alone were already surfaced to
                              # the popup UI before this field existed (see pii_detector.py's module
                              # docstring), so recording them here is always-on, not opt-in -- unlike
                              # pii_match_details, this carries no matched text, just which of
                              # pii_detector.py's ~20 patterns fired, which is what actually lets a
                              # refinement pass narrow in on which regex is noisy.
    pii_match_details: str = ""  # "" unless pii_detection.audit_match_details is turned on in
                              # settings.yaml (see pii_detector.py's is_pii_audit_match_details_
                              # enabled()) -- the opt-in PII-refinement trial capture, off by
                              # default. When on and pii_categories is non-empty: "User confirmed:
                              # details hidden" if this entry's own `decision` is NOT one of
                              # APPROVED_LIKE_DECISIONS above (rejected, denied_unattended, error --
                              # nothing was released, so nothing is recorded); otherwise
                              # "<category>: <text>" pairs (joined by "; ", one per distinct
                              # category) giving the literal matched text for a label/keyword
                              # category (e.g. "salary") or a partially redacted form for a
                              # value-bearing one (e.g. an IBAN) -- see pii_detector.py's
                              # describe_match_for_audit()/_VALUE_BEARING_CATEGORIES for exactly
                              # which categories get redacted and how.
    claude_reason: str = ""  # Claude's self-reported reason for the call, from the mandatory
                              # "reason" ToolSpec param every gated/auto tool now declares (see
                              # gate.py's reason_scope), or the "reason" param on the three
                              # privacyfence_* meta-tools for "policy_check"/
                              # "unattended_session_started"/"_ended" entries, which have no
                              # underlying gated tool call to take it from otherwise (see
                              # ipc_server.py's _audit_policy_check/_audit_unattended_session_event).
                              # Self-reported and unverified -- never treated as fact. Empty for
                              # the automatic session-end-on-disconnect path, which has no reason
                              # to attribute.
    answered_via: str = ""   # "" (the default -- today's only path, and every non-interactive
                              # decision: auto_accepted, denied_unattended, etc.) or "mobile" when
                              # a paired phone answered the approval popup instead of the desktop
                              # dialog (issue #55's mobile remote approval, Phase 1). Populated by
                              # CompositeApprovalUI racing the native and mobile-relay backends --
                              # see composite_approval_ui.py. Wiring this into gate.py's own
                              # audit() calls is left to that module's own change, not this field's
                              # addition; a schema that already has the column is what lets that
                              # land without a second migration of every existing .jsonl reader.


class AuditLogger:
    def __init__(self, log_dir: str) -> None:
        self._log_dir = Path(log_dir)
        self._log_dir.mkdir(parents=True, exist_ok=True)
        self._lock = threading.Lock()

    def record(self, entry: AuditEntry) -> None:
        week_file = self._log_dir / f"{entry.week}.jsonl"
        line = json.dumps(asdict(entry)) + "\n"
        with self._lock:
            with open(week_file, "a", encoding="utf-8") as fh:
                fh.write(line)
        logger.debug("Audit: %s %s/%s", entry.decision, entry.connector, entry.tool)

    def export_week_to_excel(self, week: str) -> str | None:
        """Export one week's .jsonl to .xlsx, overwriting any existing file.

        Callers that only want to fill in weeks that have never been
        exported should use export_all_pending() instead.
        """
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.warning("openpyxl not installed — skipping Excel audit export")
            return None

        week_file = self._log_dir / f"{week}.jsonl"
        if not week_file.exists():
            return None

        entries: list[AuditEntry] = []
        with open(week_file, encoding="utf-8") as fh:
            for line in fh:
                line = line.strip()
                if line:
                    try:
                        entries.append(AuditEntry(**json.loads(line)))
                    except Exception:
                        pass
        if not entries:
            return None

        output_path = str(self._log_dir / f"{week}.xlsx")
        wb = openpyxl.Workbook()

        # ── Main sheet ────────────────────────────────────────────────────
        ws = wb.active
        ws.title = "Decisions"

        HEADERS = [
            "Timestamp", "Week", "Connector", "Tool", "Human-Readable Name",
            "Summary", "Sender / Context", "Decision", "Auto-Accept Rule", "Latency (s)",
            "PII Detected", "PII Categories", "PII Match Details", "Claude's Reason (unverified)",
            "Answered Via",
        ]
        COL_WIDTHS = [22, 10, 12, 30, 22, 55, 30, 14, 22, 12, 12, 30, 55, 55, 14]

        hdr_font  = Font(bold=True, color="FFFFFF")
        hdr_fill  = PatternFill("solid", fgColor="2D4A6B")
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        decision_fills = {
            "approved":              PatternFill("solid", fgColor="E8F5E9"),
            "auto_accepted":         PatternFill("solid", fgColor="E3F2FD"),
            "accepted_via_accept_all": PatternFill("solid", fgColor="FFF3CD"),
            "accepted_via_temp_session": PatternFill("solid", fgColor="FFF3CD"),
            "rejected":              PatternFill("solid", fgColor="FFEBEE"),
            "denied_unattended":     PatternFill("solid", fgColor="FFD8A8"),
            "policy_check":          PatternFill("solid", fgColor="F1F3F5"),
            "rules_listed":          PatternFill("solid", fgColor="F1F3F5"),
            "rule_changed_via_bridge_proposal":   PatternFill("solid", fgColor="FFF3CD"),
            "rule_removed_via_bridge_proposal":   PatternFill("solid", fgColor="FFF3CD"),
            "grant_changed_via_bridge_proposal":  PatternFill("solid", fgColor="FFF3CD"),
            "grant_removed_via_bridge_proposal":  PatternFill("solid", fgColor="FFF3CD"),
            "bridge_proposal_no_op": PatternFill("solid", fgColor="F1F3F5"),
            "error":                 PatternFill("solid", fgColor="FF6B6B"),
        }

        ws.append(HEADERS)
        for col, _ in enumerate(HEADERS, 1):
            c = ws.cell(row=1, column=col)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = hdr_align

        for entry in entries:
            ws.append([
                entry.timestamp, entry.week, entry.connector, entry.tool,
                entry.tool_name, entry.summary, entry.sender, entry.decision,
                entry.auto_accept_rule or "", round(entry.latency_seconds, 2),
                "Yes" if entry.pii_detected else "",
                "; ".join(entry.pii_categories), entry.pii_match_details or "",
                entry.claude_reason or "", entry.answered_via or "",
            ])
            fill = decision_fills.get(entry.decision, PatternFill())
            for col in range(1, len(HEADERS) + 1):
                ws.cell(row=ws.max_row, column=col).fill = fill

        for col, width in enumerate(COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

        # ── Summary sheet ─────────────────────────────────────────────────
        ws2 = wb.create_sheet("Summary")
        ws2.append(["Metric", "Value"])
        ws2.append(["Week", week])
        ws2.append(["Total decisions", len(entries)])
        counts = Counter(e.decision for e in entries)
        ws2.append(["Approved (manual)", counts.get("approved", 0)])
        ws2.append(["Auto-accepted", counts.get("auto_accepted", 0)])
        ws2.append(["Accepted via Always allow (new rule)", counts.get("accepted_via_accept_all", 0)])
        ws2.append(["Accepted (also armed temp-accept grace window)", counts.get("accepted_via_temp_session", 0)])
        ws2.append(["Rejected", counts.get("rejected", 0)])
        ws2.append(["Denied unattended (no human asked)", counts.get("denied_unattended", 0)])
        ws2.append(["Preflight checks (privacyfence_check_policy)", counts.get("policy_check", 0)])
        ws2.append(["PII flagged (any decision)", sum(1 for e in entries if e.pii_detected)])
        ws2.append([])
        ws2.append(["By connector", ""])
        for connector, cnt in sorted(Counter(e.connector for e in entries).items()):
            ws2.append([connector, cnt])
        category_counts = Counter(cat for e in entries for cat in e.pii_categories)
        if category_counts:
            ws2.append([])
            ws2.append(["By PII category (refinement trial)", ""])
            for category, cnt in sorted(category_counts.items()):
                ws2.append([category, cnt])
        ws2.column_dimensions["A"].width = 24
        ws2.column_dimensions["B"].width = 14

        wb.save(output_path)
        logger.info("Audit Excel exported: %s (%d entries)", output_path, len(entries))
        return output_path

    def export_all_pending(self) -> None:
        """Export any week that has .jsonl but no .xlsx."""
        for jsonl in sorted(self._log_dir.glob("*.jsonl")):
            week = jsonl.stem
            xlsx = self._log_dir / f"{week}.xlsx"
            if not xlsx.exists():
                self.export_week_to_excel(week)

    def recent_entries(self, limit: int = 20) -> list[AuditEntry]:
        """Most-recent-first entries for the settings window's "Recent
        decisions" list. Reads the current ISO week's .jsonl and, if that
        alone doesn't have `limit` entries, tops up from the previous week's
        file -- no need to scan every historical file for what's meant to be
        a short "what just happened" glance, not a full audit trail browser
        (that's what the Excel export is for)."""
        weeks = [current_week(), _previous_week(current_week())]
        entries: list[AuditEntry] = []
        for week in weeks:
            week_file = self._log_dir / f"{week}.jsonl"
            if not week_file.exists():
                continue
            week_entries: list[AuditEntry] = []
            with open(week_file, encoding="utf-8") as fh:
                for line in fh:
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        week_entries.append(AuditEntry(**json.loads(line)))
                    except Exception:
                        continue
            entries.extend(reversed(week_entries))
            if len(entries) >= limit:
                break
        return entries[:limit]

    def recent_matches(self, connector: str, tool: str, summary: str, *, week: str | None = None) -> int:
        """Count prior approved-like decisions (see APPROVED_LIKE_DECISIONS)
        for the same (connector, tool, summary) in one week's log --
        defaults to the current week. The request-fingerprint feature:
        "you've approved this exact request N times this week," so a
        reviewer can spot an unusually novel request versus a routine
        repeat at a glance.

        (connector, tool, summary) is a practical proxy for "the same
        request" -- AuditEntry carries neither an operation_key nor the
        full preview dict, and summary already names the specific resource
        for most tools (e.g. "Read email: Confidential Q3 numbers", 'Read
        "Budget.xlsx"'). A coarser or finer fingerprint can replace this
        later without changing the caller-facing count semantics.
        """
        week_file = self._log_dir / f"{week or current_week()}.jsonl"
        if not week_file.exists():
            return 0
        count = 0
        with open(week_file, encoding="utf-8") as fh:
            for line in fh:
                line = line.strip()
                if not line:
                    continue
                try:
                    data = json.loads(line)
                except json.JSONDecodeError:
                    continue
                if (
                    data.get("connector") == connector
                    and data.get("tool") == tool
                    and data.get("summary") == summary
                    and data.get("decision") in APPROVED_LIKE_DECISIONS
                ):
                    count += 1
        return count


def current_week() -> str:
    iso = datetime.now(timezone.utc).isocalendar()
    return f"{iso[0]}-W{iso[1]:02d}"


def _previous_week(week: str) -> str:
    """One ISO week before `week` (e.g. "2026-W31" -> "2026-W30"), correctly
    rolling over a year boundary via datetime's own ISO calendar math rather
    than hand-rolled week-count arithmetic."""
    from datetime import timedelta

    year, week_num = week.split("-W")
    monday = datetime.fromisocalendar(int(year), int(week_num), 1)
    prev_monday = monday - timedelta(days=7)
    iso = prev_monday.isocalendar()
    return f"{iso[0]}-W{iso[1]:02d}"


_INSTANCE: AuditLogger | None = None
_LOCK = threading.Lock()


def get_audit_logger() -> AuditLogger:
    global _INSTANCE
    if _INSTANCE is None:
        with _LOCK:
            if _INSTANCE is None:
                fallback = os.path.join(os.path.expanduser("~"), ".privacyfence", "audit")
                _INSTANCE = AuditLogger(fallback)
    return _INSTANCE


def init_audit_logger(log_dir: str) -> AuditLogger:
    global _INSTANCE
    with _LOCK:
        _INSTANCE = AuditLogger(log_dir)
    return _INSTANCE
