"""Best-effort plain-text (Markdown, where the source format has real
structure worth keeping) extraction from attachment/upload content.

Feeds pii_detector.py's regex scan for formats that carry a real file's
content as opaque bytes (Gmail attachments, Drive downloads/uploads,
Confluence attachments) rather than the pre-extracted text gate.py's other
callers already have (e.g. Google Docs exports, plain-text bodies). Also
feeds the approval window's preview pane directly -- see connectors/drive.py,
gmail.py, and confluence.py, which reuse this same extracted text for both
the PII scan and a "markdown" preview_blocks entry rendered via
markdown_to_html.py. Never raises: a format this module can't parse, or
fails to parse, simply contributes no text -- the same as if the attachment
weren't there. This is a defense-in-depth heuristic layered on top of human
review (see pii_detector.py's own module docstring), not the primary
control, so silently extracting less is always the safe failure mode, never
raising into the gate path.

DOCX/PPTX/XLSX/HTML extraction keeps headings, bold/italic, bullet/numbered
lists, and tables as Markdown syntax (the same syntax html_to_text.py's
html_to_markdown() produces) rather than flattening everything to plain
prose -- a human reviewing a preview benefits from "this was a heading" or
"this was a table" surviving the conversion, even though page layout,
fonts, and other purely visual formatting is deliberately dropped. PDF has
no comparable structure available from pypdf's plain per-page
`extract_text()`, so it stays flat text.

Images are deliberately out of scope -- no OCR. A caller wanting extracted
text for image content gets an empty string here, same as any other
unsupported format.
"""
from __future__ import annotations

import io
import logging
import xml.etree.ElementTree as ET
import zipfile

import pypdf

from .html_to_text import html_to_markdown

logger = logging.getLogger(__name__)

# Cap on how much extracted text is handed to pii_detector.py's regex scan.
# Every existing caller of that scan already truncates to a much smaller
# figure for *display* purposes (Drive's 2000-char content preview), but a
# document's actual text can run far longer than what's shown in the details
# pane, so this uses a more generous cap of its own rather than inheriting
# that display-oriented number.
MAX_SCAN_CHARS = 20_000

_DOCX_MIME = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
_PPTX_MIME = "application/vnd.openxmlformats-officedocument.presentationml.presentation"
_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_ARCHIVE_MIME = "application/zip"

# Rows/columns read per XLSX sheet before truncating -- a document's actual
# grid can be far larger than anything worth dumping into a preview pane or
# a regex scan.
_XLSX_MAX_ROWS = 200
_XLSX_MAX_COLS = 20

# Archive entries listed before truncating (see _extract_archive_markdown).
_ARCHIVE_MAX_ENTRIES = 200

# MIME types extract_text() can do something with, beyond the text/* prefix.
EXTRACTABLE_MIME_TYPES = frozenset({"application/pdf", _DOCX_MIME, _PPTX_MIME, _XLSX_MIME, _ARCHIVE_MIME})


def is_prefetch_worthy(mime_type: str) -> bool:
    """Whether fetching an attachment/upload's full bytes pre-approval is
    worth it at all -- for a preview (image/*) or extracted text
    (extract_text() output, used for both the PII scan and the preview
    pane). Shared by connectors that only get one shot at the content
    before a human decides (no partial-fetch API, or an unbounded local
    disk read) and want to avoid fetching/reading types neither use case
    can do anything with.
    """
    mime_type = (mime_type or "").split(";")[0].strip().lower()
    return mime_type.startswith("image/") or mime_type.startswith("text/") or mime_type in EXTRACTABLE_MIME_TYPES


def extract_text(data: bytes, mime_type: str) -> str:
    """Best-effort text from ``data`` -- Markdown syntax where the format's
    own structure supports it (DOCX/PPTX/XLSX/HTML), plain text otherwise --
    or "" if ``mime_type`` isn't a supported format or the content fails to
    parse. Truncated to MAX_SCAN_CHARS.
    """
    if not data:
        return ""
    mime_type = (mime_type or "").split(";")[0].strip().lower()
    try:
        if mime_type == "text/html":
            text = html_to_markdown(data.decode("utf-8", errors="replace"))
        elif mime_type.startswith("text/"):
            text = data.decode("utf-8", errors="replace")
        elif mime_type == "application/pdf":
            text = _extract_pdf_text(data)
        elif mime_type == _DOCX_MIME:
            text = _extract_docx_markdown(data)
        elif mime_type == _PPTX_MIME:
            text = _extract_pptx_markdown(data)
        elif mime_type == _XLSX_MIME:
            text = _extract_xlsx_markdown(data)
        elif mime_type == _ARCHIVE_MIME:
            text = _extract_archive_markdown(data)
        else:
            return ""
    except Exception:
        logger.warning("extract_text: failed to parse %s content", mime_type, exc_info=True)
        return ""
    return text[:MAX_SCAN_CHARS]


def preview_blocks_for(details: str, extracted: str) -> list[dict] | None:
    """The WIDE right-pane ``preview_blocks`` for a "content never reaches
    Claude" download-shaped tool (drive_download_file,
    gmail_download_attachment, confluence_download_attachment,
    drive_upload_file): the boilerplate details sentence, followed by the
    file's own extracted content (see extract_text()) rendered richly via
    the "markdown" block type -- or just the details sentence alone when
    extraction found nothing, the same metadata-only fallback these tools
    have always had. None when there's neither, so a caller can leave
    gated_call's own preview_blocks default (no preview pane content)
    alone in that case.
    """
    blocks = []
    if details:
        blocks.append({"type": "text", "text": details})
    if extracted:
        blocks.append({"type": "markdown", "text": extracted})
    return blocks or None


def _extract_pdf_text(data: bytes) -> str:
    """Page text joined in document order, stopping as soon as the running
    total passes MAX_SCAN_CHARS rather than always walking every page --
    extract_text()'s own truncation to that same cap happens after this
    returns, so a large multi-hundred-page PDF would otherwise pay for
    pages whose text never survives the truncation anyway, in the
    synchronous approval-gate path where a human is waiting on the popup.
    """
    reader = pypdf.PdfReader(io.BytesIO(data))
    parts = []
    length = 0
    for page in reader.pages:
        text = page.extract_text() or ""
        parts.append(text)
        length += len(text)
        if length >= MAX_SCAN_CHARS:
            break
    return "\n".join(parts)


def _local(tag: str) -> str:
    """Element local name, stripping ElementTree's ``{namespace}`` prefix --
    the same namespace-agnostic trick this module has always used (matching
    on tag suffix), extended to attribute-bearing lookups below. Lets one
    walk work across DOCX/PPTX regardless of exactly which namespace URI a
    producer bound a prefix to."""
    return tag.rsplit("}", 1)[-1]


def _attr_val(elem) -> str:
    """The ``val`` attribute of ``elem`` (WordprocessingML's boolean/style
    properties always carry it as ``w:val``), regardless of namespace
    prefix. "" if absent."""
    for key, value in elem.attrib.items():
        if key == "val" or key.endswith("}val"):
            return value
    return ""


def _bool_prop(elem) -> bool:
    """WordprocessingML boolean run/paragraph properties (``<w:b/>``,
    ``<w:i/>``) default to true when present with no ``val`` at all --
    only an explicit ``val="false"``/``"0"`` turns them off."""
    return _attr_val(elem) not in ("false", "0")


def _join_markdown_lines(lines: list[str]) -> str:
    """Join extracted heading/bullet/paragraph lines with a blank line
    between distinct blocks, except a single newline between two
    consecutive bullet lines -- markdown_to_html.py groups a run of
    "- "-prefixed lines with no blank line between them into one <ul>
    (see its own _render_block), so keeping a list's own items
    unseparated is what makes consecutive bullets render as one list
    instead of several single-item ones."""
    if not lines:
        return ""
    parts = [lines[0]]
    for prev, line in zip(lines, lines[1:]):
        parts.append("\n" if prev.startswith("- ") and line.startswith("- ") else "\n\n")
        parts.append(line)
    return "".join(parts)


# ---------------------------------------------------------------------------- #
# DOCX -- Markdown, one heading/bullet/paragraph line per <w:p>
# ---------------------------------------------------------------------------- #

def _extract_docx_markdown(data: bytes) -> str:
    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        if "word/document.xml" not in zf.namelist():
            return ""
        root = ET.fromstring(zf.read("word/document.xml"))
    raw_lines = (_docx_paragraph_markdown(p) for p in root.iter() if _local(p.tag) == "p")
    return _join_markdown_lines([line for line in raw_lines if line])


def _docx_paragraph_markdown(p) -> str:
    heading_level = 0
    is_bullet = False
    for child in p.iter():
        name = _local(child.tag)
        if name == "pStyle":
            style = _attr_val(child)
            if style == "Title":
                heading_level = 1
            elif style.startswith("Heading"):
                digits = "".join(c for c in style if c.isdigit())
                heading_level = min(int(digits), 6) if digits else 1
        elif name == "numPr":
            is_bullet = True
    text = _docx_runs_markdown(p).strip()
    if not text:
        return ""
    if heading_level:
        return f"{'#' * heading_level} {text}"
    if is_bullet:
        return f"- {text}"
    return text


def _docx_runs_markdown(p) -> str:
    parts = []
    for r in p.iter():
        if _local(r.tag) != "r":
            continue
        bold = italic = False
        run_text_parts = []
        for child in r:
            name = _local(child.tag)
            if name == "rPr":
                for prop in child:
                    prop_name = _local(prop.tag)
                    if prop_name == "b" and _bool_prop(prop):
                        bold = True
                    elif prop_name == "i" and _bool_prop(prop):
                        italic = True
            elif name == "t" and child.text:
                run_text_parts.append(child.text)
        run_text = "".join(run_text_parts)
        if not run_text:
            continue
        if bold:
            run_text = f"**{run_text}**"
        if italic:
            run_text = f"*{run_text}*"
        parts.append(run_text)
    return "".join(parts)


# ---------------------------------------------------------------------------- #
# PPTX -- Markdown, "## Slide N" heading per slide followed by that slide's
# own paragraph/bullet lines
# ---------------------------------------------------------------------------- #

def _pptx_slide_paths(data: bytes) -> list[str]:
    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        return sorted(
            n for n in zf.namelist()
            if n.startswith("ppt/slides/slide") and n.endswith(".xml")
        )


def _extract_pptx_markdown(data: bytes) -> str:
    paths = _pptx_slide_paths(data)
    if not paths:
        return ""
    sections = []
    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        for i, path in enumerate(paths, start=1):
            root = ET.fromstring(zf.read(path))
            raw_lines = (_pptx_paragraph_markdown(p) for p in root.iter() if _local(p.tag) == "p")
            body = _join_markdown_lines([line for line in raw_lines if line])
            sections.append(f"## Slide {i}\n\n{body}" if body else f"## Slide {i}")
    return "\n\n".join(sections)


def _pptx_paragraph_markdown(p) -> str:
    bullet = False
    parts = []
    for child in p:
        name = _local(child.tag)
        if name == "pPr":
            bullet = any(_local(prop.tag) in ("buChar", "buAutoNum") for prop in child)
        elif name == "r":
            run_text_parts = []
            bold = italic = False
            for rchild in child:
                rname = _local(rchild.tag)
                if rname == "rPr":
                    bold = rchild.attrib.get("b") in ("1", "true")
                    italic = rchild.attrib.get("i") in ("1", "true")
                elif rname == "t" and rchild.text:
                    run_text_parts.append(rchild.text)
            run_text = "".join(run_text_parts)
            if run_text:
                if bold:
                    run_text = f"**{run_text}**"
                if italic:
                    run_text = f"*{run_text}*"
                parts.append(run_text)
    text = "".join(parts).strip()
    if not text:
        return ""
    return f"- {text}" if bullet else text


# ---------------------------------------------------------------------------- #
# XLSX -- Markdown table per sheet, via openpyxl (already a hard dependency,
# see pyproject.toml -- imported lazily here to match audit_log.py's own
# established "degrade gracefully if it's somehow missing" precedent for
# this specific package rather than a hard top-level import)
# ---------------------------------------------------------------------------- #

def _extract_xlsx_markdown(data: bytes) -> str:
    try:
        import openpyxl
    except ImportError:
        logger.warning("extract_text: openpyxl not installed -- skipping XLSX extraction")
        return ""
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        sections = [s for s in (_xlsx_sheet_markdown(ws) for ws in wb.worksheets) if s]
    finally:
        wb.close()
    return "\n\n".join(sections)


def _xlsx_sheet_markdown(ws) -> str:
    max_col = min(ws.max_column, _XLSX_MAX_COLS) if ws.max_column else _XLSX_MAX_COLS
    rows = []
    for row in ws.iter_rows(max_row=_XLSX_MAX_ROWS, max_col=max_col, values_only=True):
        cells = ["" if v is None else str(v) for v in row]
        if any(cell for cell in cells):
            rows.append(cells)
    if not rows:
        return ""
    header, *body = rows
    lines = [_row_markdown(header), _row_markdown(["---"] * len(header))]
    lines.extend(_row_markdown(row) for row in body)
    section = f"## {ws.title}\n\n" + "\n".join(lines)
    if ws.max_row and ws.max_row > _XLSX_MAX_ROWS:
        section += f"\n\n*(showing the first {_XLSX_MAX_ROWS} rows of {ws.max_row})*"
    return section


def _row_markdown(cells: list[str]) -> str:
    return "| " + " | ".join(cell.replace("|", "\\|") for cell in cells) + " |"


# ---------------------------------------------------------------------------- #
# Archives -- not text content, but a member-file listing is far more useful
# in a preview than nothing at all
# ---------------------------------------------------------------------------- #

def _extract_archive_markdown(data: bytes) -> str:
    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        entries = [info for info in zf.infolist() if not info.is_dir()][:_ARCHIVE_MAX_ENTRIES]
    if not entries:
        return ""
    lines = [_row_markdown(["Name", "Size"]), _row_markdown(["---", "---"])]
    lines.extend(_row_markdown([info.filename, f"{info.file_size:,} bytes"]) for info in entries)
    return "\n".join(lines)
